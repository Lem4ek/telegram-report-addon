import asyncio
from datetime import datetime, timedelta, date as _date
import os

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
from telegram import (
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
    KeyboardButton,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
    Update,
)
from telegram.ext import ApplicationBuilder, MessageHandler, filters, CommandHandler, CallbackQueryHandler

from parser import parse_message
from data_utils import save_entry, generate_stats, get_csv_file, get_month_file_str

# ────────────────────────────────────────────────
# Конфигурация
# ────────────────────────────────────────────────
TOKEN = os.getenv("TELEGRAM_TOKEN")

def _parse_ids(s: str) -> set[int]:
    ids = set()
    for part in (s or "").replace(";", ",").split(","):
        part = part.strip()
        if not part:
            continue
        try:
            ids.add(int(part))
        except ValueError:
            pass
    return ids

# Пример: ALLOWED_USER_IDS="1198365511,508532161"
ALLOWED_USER_IDS: set[int] = _parse_ids(os.getenv("ALLOWED_USER_IDS", ""))
if not ALLOWED_USER_IDS:
    print("[WARN] ALLOWED_USER_IDS не задан — доступ будет закрыт всем пользователям.")
else:
    print(f"[INFO] Разрешенные ID: {sorted(ALLOWED_USER_IDS)}")

user_stats = {}
current_month = datetime.now().month
pending_updates = {}

SAVE_DELAY = timedelta(minutes=2)  # задержка перед записью/отправкой


# ────────────────────────────────────────────────
# Утилиты
# ────────────────────────────────────────────────
def safe_float(value):
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0


def is_allowed(update: Update):
    try:
        user_id = update.effective_user.id
    except Exception:
        return False
    return user_id in ALLOWED_USER_IDS


def is_valid_report(text: str) -> bool:
    """Фильтр отчёта: требуем присутствие ключевых слов."""
    t = text.lower()
    groups = [
        ["паков", "паки", "упаков"],
        ["вес"],
        ["отход"],
        ["пакетосвар"],
        ["экструз", "экструд"],
    ]
    return all(any(v in t for v in grp) for grp in groups)


def load_stats_from_excel():
    """Загружает статистику из текущего Excel в user_stats при старте."""
    file_path = get_csv_file()
    if not os.path.exists(file_path):
        return

    user_stats.clear()
    wb = load_workbook(file_path)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        date, user, pakov, ves, paket, flexa, extru, itogo = row
        if not user:
            continue

        if user not in user_stats:
            user_stats[user] = {
                'Паков': 0.0, 'Вес': 0.0, 'Пакетосварка': 0.0,
                'Флекса': 0.0, 'Экструзия': 0.0, 'Итого': 0.0, 'Смен': 0
            }

        # инкремент смен на каждую строку
        user_stats[user]['Смен'] += 1

        # накопительные метрики
        user_stats[user]['Паков']        += safe_float(pakov)
        user_stats[user]['Вес']          += safe_float(ves)
        user_stats[user]['Пакетосварка'] += safe_float(paket)
        user_stats[user]['Флекса']       += safe_float(flexa)
        user_stats[user]['Экструзия']    += safe_float(extru)
        user_stats[user]['Итого']        += safe_float(itogo)


# ────────────────────────────────────────────────
# Кнопки
# ────────────────────────────────────────────────
def build_main_keyboard() -> ReplyKeyboardMarkup:
    kb = [
        [KeyboardButton("/graf"), KeyboardButton("/csv")],
        [KeyboardButton("/stats"), KeyboardButton("/myid")],
        [KeyboardButton("📥 Импорт месяца…")],
    ]
    return ReplyKeyboardMarkup(kb, resize_keyboard=True, one_time_keyboard=False)

def prev_month_str(today: _date | None = None) -> str:
    if today is None:
        today = _date.today()
    y, m = today.year, today.month
    if m == 1:
        return f"{y-1}-12"
    return f"{y}-{m-1:02d}"

def cur_month_str(today: _date | None = None) -> str:
    if today is None:
        today = _date.today()
    return f"{today.year}-{today.month:02d}"


# ────────────────────────────────────────────────
# Сохранение отчёта с задержкой (debounce)
# ────────────────────────────────────────────────
async def delayed_save(message_id):
    try:
        await asyncio.sleep(SAVE_DELAY.total_seconds())
        if message_id not in pending_updates:
            return

        data = pending_updates.pop(message_id)
        bot = data["context"].bot
        chat_id = data["chat_id"]
        username = data["user"]
        values = data["values"]

        # 1) Сохраняем в Excel (ротация по месяцу внутри save_entry)
        save_entry(data["time"], username, values)

        # 2) Обновляем оперативную статистику
        user_stats.setdefault(username, {
            'Паков': 0.0, 'Вес': 0.0, 'Пакетосварка': 0.0,
            'Флекса': 0.0, 'Экструзия': 0.0, 'Итого': 0.0, 'Смен': 0
        })
        for k, v in values.items():
            if k in user_stats[username] and isinstance(v, (int, float)):
                user_stats[username][k] += v
        user_stats[username]['Смен'] += 1

        total_pakov_all = sum(u.get('Паков', 0.0) for u in user_stats.values())
        total_ves_all = sum(u.get('Вес', 0.0) for u in user_stats.values())

        # 3) Сообщение в чат
        report = f"""
📦 Отчёт за смену:

📦  Паков: {values.get('Паков', 0.0):.2f} шт
⚖️ Вес: {values.get('Вес', 0.0):.2f} кг

♻️ Отходы:
🛍️ Пакетосварка: {values.get('Пакетосварка', 0.0):.2f} кг
🎨 Флекса: {values.get('Флекса', 0.0):.2f} кг
🧵 Экструзия: {values.get('Экструзия', 0.0):.2f} кг

🧾 Итого отходов: {values.get('Итого', 0.0):.2f} кг

📊 Всего продукции за период: {total_pаков_all:.2f} паков / {total_ves_all:.2f} кг
""".strip()

        await bot.send_message(chat_id=chat_id, text=report)

    except Exception as e:
        import traceback
        print("delayed_save error:", e)
        print(traceback.format_exc())
        try:
            await data["context"].bot.send_message(
                chat_id=data.get("chat_id"),
                text=f"✅ Отчёт сохранён, но ошибка при отправке сообщения: {e}"
            )
        except Exception:
            pass


# ────────────────────────────────────────────────
# Хэндлеры сообщений
# ────────────────────────────────────────────────
async def handle_message(update, context):
    global current_month
    month_now = datetime.now().month
    if month_now != current_month:
        user_stats.clear()
        pending_updates.clear()
        current_month = month_now

    if not update.message or not update.message.text:
        return

    username = update.effective_user.first_name
    text = update.message.text.strip()

    # быстрые кнопки: "Импорт месяца…"
    if text == "📥 Импорт месяца…":
        # показать inline-клавиатуру с готовыми месяцами
        pm = prev_month_str()
        cm = cur_month_str()
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton(f"⬅️ {pm}", callback_data=f"import_month:{pm}"),
             InlineKeyboardButton(f"📅 {cm}", callback_data=f"import_month:{cm}")],
        ])
        await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text="Выберите месяц или введите команду вручную: `/import YYYY-MM`",
            reply_markup=kb,
            parse_mode="Markdown"
        )
        return

    # обычный отчёт
    if not is_valid_report(text):
        return

    values = parse_message(text)
    if not values:
        return

    # минимум 3 непустых поля
    if sum(1 for v in values.values() if v not in (0, "", None)) < 3:
        return

    for key in ["Паков", "Вес", "Пакетосварка", "Флекса", "Экструзия"]:
        values.setdefault(key, 0.0)

    values["Итого"] = (
        safe_float(values.get("Пакетосварка", 0))
        + safe_float(values.get("Флекса", 0))
        + safe_float(values.get("Экструзия", 0))
    )

    message_id = update.message.message_id
    pending_updates[message_id] = {
        "user": username,
        "values": values,
        "time": datetime.now(),
        "chat_id": update.effective_chat.id,
        "context": context
    }

    asyncio.create_task(delayed_save(message_id))


async def handle_edited_message(update, context):
    if not update.edited_message or not update.edited_message.text:
        return
    message_id = update.edited_message.message_id
    if message_id not in pending_updates:
        return

    text = update.edited_message.text
    if not is_valid_report(text):
        return

    values = parse_message(text)
    if sum(1 for v in values.values() if v not in (0, "", None)) < 3:
        return

    for key in ["Паков", "Вес", "Пакетосварка", "Флекса", "Экструзия"]:
        values.setdefault(key, 0.0)

    values["Итого"] = (
        safe_float(values.get("Пакетосварка", 0))
        + safe_float(values.get("Флекса", 0))
        + safe_float(values.get("Экструзия", 0))
    )

    pending_updates[message_id]["values"] = values
    pending_updates[message_id]["time"] = datetime.now()


# ────────────────────────────────────────────────
# Команды
# ────────────────────────────────────────────────
async def cmd_start_menu(update, context):
    if not is_allowed(update):
        await context.bot.send_message(chat_id=update.effective_chat.id, text="⛔ Нет доступа.")
        return
    await context.bot.send_message(
        chat_id=update.effective_chat.id,
        text="Готов к работе. Выберите действие на клавиатуре ниже:",
        reply_markup=build_main_keyboard()
    )

async def cmd_hide(update, context):
    if not is_allowed(update):
        return
    await context.bot.send_message(
        chat_id=update.effective_chat.id,
        text="Клавиатура скрыта. /menu чтобы вернуть.",
        reply_markup=ReplyKeyboardRemove()
    )

async def cmd_csv(update, context):
    if not is_allowed(update):
        await context.bot.send_message(chat_id=update.effective_chat.id, text="⛔ Нет доступа.")
        return

    # если передан аргумент формата YYYY-MM — шлём файл за этот месяц
    if context.args:
        ym = context.args[0].strip()
        from datetime import datetime as _dt
        try:
            _dt.strptime(ym, "%Y-%m")
        except Exception:
            await context.bot.send_message(
                chat_id=update.effective_chat.id,
                text="Неверный формат. Использование: /csv или /csv YYYY-MM (например, /csv 2025-07)"
            )
            return

        try:
            from data_utils import get_month_file_str
            file_path = get_month_file_str(ym)
        except Exception:
            file_path = os.path.join("/config/bnk_bot/data", f"{ym}.xlsx")

        if not os.path.exists(file_path):
            await context.bot.send_message(chat_id=update.effective_chat.id, text=f"Файл за {ym} не найден.")
            return

        await context.bot.send_document(
            chat_id=update.effective_chat.id,
            document=open(file_path, "rb"),
            filename=f"BNK_{ym}.xlsx"
        )
        return

    # без аргументов — шлём текущий месяц
    file_path = get_csv_file()
    if not os.path.exists(file_path):
        await context.bot.send_message(chat_id=update.effective_chat.id, text="❌ Файл данных не найден.")
        return

    await context.bot.send_document(
        chat_id=update.effective_chat.id,
        document=open(file_path, 'rb'),
        filename=f"BNK_{datetime.now():%Y-%m}.xlsx"
    )


async def cmd_stats(update, context):
    if not is_allowed(update):
        await context.bot.send_message(chat_id=update.effective_chat.id, text="⛔ Нет доступа.")
        return
    await context.bot.send_message(chat_id=update.effective_chat.id, text=generate_stats(user_stats))

async def cmd_reset(update, context):
    if not is_allowed(update):
        await context.bot.send_message(chat_id=update.effective_chat.id, text="⛔ Нет доступа.")
        return
    user_stats.clear()
    pending_updates.clear()
    await context.bot.send_message(chat_id=update.effective_chat.id, text="♻️ Статистика и буфер сброшены!")

async def cmd_myid(update, context):
    if update.message.chat.type != "private":
        await context.bot.send_message(chat_id=update.effective_chat.id, text="ℹ️ Запросите ID в личке.")
        return
    await context.bot.send_message(
        chat_id=update.effective_chat.id,
        text=f"🆔 Ваш Telegram ID: `{update.effective_user.id}`",
        parse_mode="Markdown"
    )

# /import YYYY-MM — выдать файл за месяц
async def cmd_import_month(update, context):
    if not is_allowed(update):
        await context.bot.send_message(chat_id=update.effective_chat.id, text="⛔ Нет доступа.")
        return

    if not context.args:
        await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text="Использование: /import YYYY-MM (например, /import 2024-09)"
        )
        return

    ym = context.args[0].strip()
    from datetime import datetime as _dt
    try:
        _dt.strptime(ym, "%Y-%m")
    except Exception:
        await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text="Неверный формат. Нужно YYYY-MM, например 2024-09."
        )
        return

    file_path = get_month_file_str(ym)
    if not os.path.exists(file_path):
        await context.bot.send_message(chat_id=update.effective_chat.id, text=f"Файл за {ym} не найден.")
        return

    await context.bot.send_document(
        chat_id=update.effective_chat.id,
        document=open(file_path, "rb"),
        filename=f"BNK_{ym}.xlsx"
    )

# Старый импорт: отправить Excel с подписью /import
async def cmd_import(update, context):
    if not is_allowed(update):
        await context.bot.send_message(chat_id=update.effective_chat.id, text="⛔ Нет доступа.")
        return
    from pathlib import Path
    current_file = Path(get_csv_file())
    if current_file.exists():
        current_file.unlink()
        user_stats.clear()

    if not update.message.document:
        await context.bot.send_message(chat_id=update.effective_chat.id, text="📄 Пожалуйста, отправьте Excel-файл.")
        return

    file = await update.message.document.get_file()
    file_path = f"/tmp/imported.xlsx"
    await file.download_to_drive(file_path)

    try:
        wb = load_workbook(file_path)
        ws = wb.active
        imported = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            date_str, user, pakov, ves, paket, flexa, extru, itogo = row
            if not user:
                continue

            try:
                if isinstance(date_str, str):
                    date = datetime.strptime(date_str, "%Y-%m-%d %H:%M")
                else:
                    date = date_str
            except Exception:
                date = datetime.now()

            values = {
                "Паков": pakov or 0,
                "Вес": ves or 0,
                "Пакетосварка": paket or 0,
                "Флекса": flexa or 0,
                "Экструзия": extru or 0,
                "Итого": itogo or 0
            }

            save_entry(date, user, values)

            if user not in user_stats:
                user_stats[user] = {
                    'Паков': 0.0, 'Вес': 0.0, 'Пакетосварка': 0.0,
                    'Флекса': 0.0, 'Экструзия': 0.0, 'Итого': 0.0, 'Смен': 0
                }

            user_stats[user]['Паков'] += float(values["Паков"])
            user_stats[user]['Вес'] += float(values["Вес"])
            user_stats[user]['Пакетосварка'] += float(values["Пакетосварка"])
            user_stats[user]['Флекса'] += float(values["Флекса"])
            user_stats[user]['Экструзия'] += float(values["Экструзия"])
            user_stats[user]['Итого'] += float(values["Итого"])
            user_stats[user]['Смен'] += 1

            imported += 1

        await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text=f"✅ Импорт завершён. Загружено и сохранено записей: {imported}"
        )

    except Exception as e:
        await context.bot.send_message(chat_id=update.effective_chat.id, text=f"❌ Ошибка при импорте: {e}")


# ────────────────────────────────────────────────
# Inline callback (импорт месяца по кнопке)
# ────────────────────────────────────────────────
async def on_callback(update, context):
    if not is_allowed(update):
        return
    q = update.callback_query
    if not q or not q.data:
        return
    await q.answer()
    if q.data.startswith("import_month:"):
        ym = q.data.split(":", 1)[1]
        file_path = get_month_file_str(ym)
        if not os.path.exists(file_path):
            await q.edit_message_text(f"Файл за {ym} не найден.")
            return
        await context.bot.send_document(
            chat_id=update.effective_chat.id,
            document=open(file_path, "rb"),
            filename=f"BNK_{ym}.xlsx"
        )

# ────────────────────────────────────────────────
# Графики
# ────────────────────────────────────────────────
async def cmd_graf(update, context):
    if not is_allowed(update):
        await context.bot.send_message(chat_id=update.effective_chat.id, text="⛔ Нет доступа.")
        return

    file_path = get_csv_file()
    if not os.path.exists(file_path):
        await context.bot.send_message(chat_id=update.effective_chat.id, text="❌ Файл данных не найден.")
        return

    df = pd.read_excel(file_path)
    if df.empty:
        await context.bot.send_message(chat_id=update.effective_chat.id, text="❌ В файле нет данных.")
        return

    df.columns = ["Дата", "Имя", "Паков", "Вес", "Пакетосварка", "Флекса", "Экструзия", "Итого"]
    try:
        df["Дата"] = pd.to_datetime(df["Дата"])
    except Exception:
        pass

    # ГРАФИК 1 — линия по дням
    daily = df.groupby(df["Дата"].dt.date).agg({"Вес": "sum", "Итого": "sum"}).reset_index()

    fig, ax = plt.subplots()
    ax.plot(daily["Дата"], daily["Вес"], marker="o", label="Вес (кг)")
    ax.plot(daily["Дата"], daily["Итого"], marker="o", label="Отходы (кг)")

    ax.set_title("Производство и отходы по дням")
    ax.set_xlabel("Дата")
    ax.set_ylabel("Кг")
    ax.legend()
    ax.grid(True, alpha=0.25)
    fig.autofmt_xdate()

    ymin, ymax = ax.get_ylim()
    dy = max(1, (ymax - ymin) * 0.02)
    for x, y in zip(daily["Дата"], daily["Вес"]):
        ax.text(x, y + dy, f"{y:.0f}", ha="center", va="bottom", fontsize=8)
    for x, y in zip(daily["Дата"], daily["Итого"]):
        ax.text(x, y + dy, f"{y:.0f}", ha="center", va="bottom", fontsize=8)

    fig.tight_layout()
    img1 = "/tmp/graf1_daily.png"
    fig.savefig(img1)
    plt.close(fig)
    await context.bot.send_photo(chat_id=update.effective_chat.id, photo=open(img1, "rb"))

    # ГРАФИК 2 — топ по весу
    top_users = (
        df.groupby("Имя")["Вес"]
        .sum()
        .sort_values(ascending=False)
        .head(10)
    )
    plt.figure()
    plt.bar(top_users.index, top_users.values)
    plt.title("ТОП производители по весу (кг)")
    plt.xlabel("Пользователь")
    plt.ylabel("Кг")
    plt.xticks(rotation=45, ha="right")
    for i, v in enumerate(top_users.values):
        plt.text(i, v, f"{v:.0f}", ha="center", va="bottom")

    plt.tight_layout()
    img2 = "/tmp/graf2_top_weight.png"
    plt.savefig(img2)
    plt.close()
    await context.bot.send_photo(chat_id=update.effective_chat.id, photo=open(img2, "rb"))

    # ГРАФИК 3 — доля отходов
    total_weight = float(df["Вес"].sum())
    total_waste = float(df["Итого"].sum())
    good = max(total_weight - total_waste, 0)

    plt.figure()
    plt.pie([good, total_waste],
            labels=["Продукция", "Отходы"],
            autopct="%1.1f%%",
            startangle=90)
    plt.axis("equal")
    plt.title("Доля отходов в общей массе")
    img3 = "/tmp/graf3_share.png"
    plt.savefig(img3)
    plt.close()
    await context.bot.send_photo(chat_id=update.effective_chat.id, photo=open(img3, "rb"))

    # ГРАФИК 4 — топ по браку
    agg = df.groupby("Имя")[["Итого"]].sum().reset_index()
    top_kg = agg.sort_values("Итого", ascending=False).head(10)

    plt.figure()
    plt.bar(top_kg["Имя"], top_kg["Итого"])
    plt.title("Топ по браку (кг)")
    plt.xlabel("Пользователь")
    plt.ylabel("Брак, кг")
    plt.xticks(rotation=45, ha="right")
    for i, v in enumerate(top_kg["Итого"]):
        plt.text(i, v, f"{v:.0f}", ha="center", va="bottom")

    plt.tight_layout()
    img4 = "/tmp/graf4_defect_kg.png"
    plt.savefig(img4)
    plt.close()
    await context.bot.send_photo(chat_id=update.effective_chat.id, photo=open(img4, "rb"))


# ────────────────────────────────────────────────
# Точка входа
# ────────────────────────────────────────────────
def main():
    if not TOKEN:
        raise ValueError("TELEGRAM_TOKEN env variable is required")

    load_stats_from_excel()

    app = ApplicationBuilder().token(TOKEN).build()

    # меню и клавиатура
    app.add_handler(CommandHandler("start", cmd_start_menu))
    app.add_handler(CommandHandler("menu", cmd_start_menu))
    app.add_handler(CommandHandler("hide", cmd_hide))

    app.add_handler(CommandHandler("csv", cmd_csv))
    app.add_handler(CommandHandler("stats", cmd_stats))
    app.add_handler(CommandHandler("reset", cmd_reset))
    app.add_handler(CommandHandler("myid", cmd_myid))

    # новая команда: /import YYYY-MM -> отдать файл месяца
    app.add_handler(CommandHandler("import", cmd_import_month))

    # старый импорт: отправьте Excel и в подписи укажите /import
    app.add_handler(MessageHandler(filters.Document.ALL & filters.Caption("/import"), cmd_import))

    # inline callback (кнопки импорта месяца)
    app.add_handler(CallbackQueryHandler(on_callback))

    app.add_handler(CommandHandler("graf", cmd_graf))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    app.add_handler(MessageHandler(filters.UpdateType.EDITED_MESSAGE, handle_edited_message))

    app.run_polling()


if __name__ == "__main__":
    main()
