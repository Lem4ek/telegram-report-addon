import os
from datetime import datetime

from openpyxl import Workbook, load_workbook

# ──────────────────────────────────────────────────────────────────────────────
# Папки для данных и графиков
# ──────────────────────────────────────────────────────────────────────────────
DATA_DIR = "/config/bnk_bot/data"
GRAPHS_DIR = "/config/bnk_bot/graphs"
os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(GRAPHS_DIR, exist_ok=True)


# ──────────────────────────────────────────────────────────────────────────────
# Пути к файлам данных (с ротацией по месяцам)
# ──────────────────────────────────────────────────────────────────────────────
def get_file_path(dt=None) -> str:
    """
    Путь к Excel-файлу месяца по дате dt (если dt=None — текущая дата).
    Формат: /config/bnk_bot/data/YYYY-MM.xlsx
    """
    if dt is None:
        dt = datetime.now()
    return os.path.join(DATA_DIR, f"{dt.strftime('%Y-%m')}.xlsx")


def get_csv_file() -> str:
    """
    Совместимость с существующим кодом: вернуть путь к файлу ТЕКУЩЕГО месяца.
    """
    return get_file_path(datetime.now())


def get_month_file_str(ym: str) -> str:
    """
    Вспомогательно: путь к файлу месяца по строке 'YYYY-MM' (без проверки существования).
    """
    return os.path.join(DATA_DIR, f"{ym}.xlsx")


# ──────────────────────────────────────────────────────────────────────────────
# Запись данных
# ──────────────────────────────────────────────────────────────────────────────
def save_entry(date: datetime, user: str, values: dict):
    """
    Сохраняем одну строку в Excel МЕСЯЦА, соответствующего дате записи.
    Колонки: Дата | Имя | Паков | Вес | Пакетосварка | Флекса | Экструзия | Итого
    """
    file_path = get_file_path(date)  # ключевая строка: ротация по месяцу записи

    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Дата", "Имя", "Паков", "Вес", "Пакетосварка", "Флекса", "Экструзия", "Итого"])

    ws.append([
        date.strftime('%Y-%m-%d %H:%M'),
        user,
        float(values.get("Паков", 0) or 0),
        float(values.get("Вес", 0) or 0),
        float(values.get("Пакетосварка", 0) or 0),
        float(values.get("Флекса", 0) or 0),
        float(values.get("Экструзия", 0) or 0),
        float(values.get("Итого", 0) or 0),
    ])

    wb.save(file_path)


# ──────────────────────────────────────────────────────────────────────────────
# Текстовая статистика /stats
# ──────────────────────────────────────────────────────────────────────────────
def generate_stats(stats: dict) -> str:
    """
    Текстовая сводка /stats по пользователям.
    Ожидается структура:
      user_stats[user] = {
          'Паков': float, 'Вес': float, 'Пакетосварка': float, 'Флекса': float,
          'Экструзия': float, 'Итого': float, 'Смен': int
      }
    """
    if not stats:
        return "📊 Статистика пуста."

    lines = ["📊 Статистика по пользователям:"]
    for user, data in stats.items():
        lines.append(
            f"{user}:\n"
            f"  📦 Паков: {float(data.get('Паков', 0)):.2f} шт\n"
            f"  ⚖️ Вес: {float(data.get('Вес', 0)):.2f} кг\n"
            f"  🛍️ Пакетосварка: {float(data.get('Пакетосварка', 0)):.2f} кг\n"
            f"  🎨 Флекса: {float(data.get('Флекса', 0)):.2f} кг\n"
            f"  🧵 Экструзия: {float(data.get('Экструзия', 0)):.2f} кг\n"
            f"  🧾 Итого отходов: {float(data.get('Итого', 0)):.2f} кг\n"
            f"  🗓 Смен: {int(data.get('Смен', 0))}"
        )
    return "\n\n".join(lines)


# ──────────────────────────────────────────────────────────────────────────────
# Генерация простых графиков по пользователям (опционально)
# ──────────────────────────────────────────────────────────────────────────────
def _ensure_matplotlib() -> bool:
    """
    Ленивая загрузка matplotlib, чтобы не падать при импорте,
    если модуль не установлен (но он должен быть в requirements.txt).
    """
    try:
        import matplotlib  # noqa: F401
        import matplotlib.pyplot as plt  # noqa: F401
        return True
    except Exception as e:
        print(f"[WARN] matplotlib не найден или не загрузился: {e}")
        return False


def generate_graphs(user_stats: dict):
    """
    Строит 3 графика (бар-чарты) по пользователям и возвращает список путей к PNG:
      1) Паков по пользователям
      2) Вес по пользователям
      3) Итого отходов по пользователям
    Примечание: в основном боте графики строятся в /graf внутри main.py,
    эта функция — чисто вспомогательная/дополнительная.
    """
    if not _ensure_matplotlib():
        return []

    import matplotlib.pyplot as plt  # требование: не задавать стили и цвета

    users = list(user_stats.keys())
    if not users:
        return []

    pakov_vals = [float(user_stats[u].get('Паков', 0.0)) for u in users]
    ves_vals   = [float(user_stats[u].get('Вес', 0.0)) for u in users]
    waste_vals = [float(user_stats[u].get('Итого', 0.0)) for u in users]

    out_files = []

    # 1) Паков
    fig = plt.figure()
    plt.bar(users, pakov_vals)
    plt.title("Паков по пользователям")
    plt.xlabel("Пользователи")
    plt.ylabel("Паков, шт")
    plt.xticks(rotation=30, ha='right')
    path1 = os.path.join(GRAPHS_DIR, "pakov_by_user.png")
    fig.tight_layout()
    fig.savefig(path1)
    plt.close(fig)
    out_files.append(path1)

    # 2) Вес
    fig = plt.figure()
    plt.bar(users, ves_vals)
    plt.title("Вес по пользователям")
    plt.xlabel("Пользователи")
    plt.ylabel("Вес, кг")
    plt.xticks(rotation=30, ha='right')
    path2 = os.path.join(GRAPHS_DIR, "ves_by_user.png")
    fig.tight_layout()
    fig.savefig(path2)
    plt.close(fig)
    out_files.append(path2)

    # 3) Итого отходов
    fig = plt.figure()
    plt.bar(users, waste_vals)
    plt.title("Итого отходов по пользователям")
    plt.xlabel("Пользователи")
    plt.ylabel("Отходы, кг")
    plt.xticks(rotation=30, ha='right')
    path3 = os.path.join(GRAPHS_DIR, "waste_by_user.png")
    fig.tight_layout()
    fig.savefig(path3)
    plt.close(fig)
    out_files.append(path3)

    return out_files
