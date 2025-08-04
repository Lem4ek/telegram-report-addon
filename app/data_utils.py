 import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

def get_excel_path(folder: str) -> str:
    os.makedirs(folder, exist_ok=True)
    filename = f"{datetime.now().strftime('%Y-%m')}.xlsx"
    return os.path.join(folder, filename)

def clean_extrusion(raw: str) -> str:
    if not raw:
        return ""
    return raw.lower().replace("экструзия", "").strip()

def save_data(data: dict, user: str, user_id: int, folder: str):
    path = get_excel_path(folder)
    file_exists = os.path.isfile(path)

    if file_exists:
        wb = load_workbook(path)
    else:
        wb = Workbook()

    sheet = wb.active
    sheet.title = "Данные"

    if not file_exists:
        headers = ["Дата", "Пользователь", "ID", "Паков", "Вес", "Пакетосварка", "Флекса", "Экструзия", "Итого"]
        sheet.append(headers)

    # Удаляем старую строку "ИТОГО", если она есть
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        if row[0].value and str(row[0].value).lower() == "итого":
            sheet.delete_rows(row[0].row, 1)
            break

    sheet.append([
        datetime.now().strftime('%d.%m.%Y'),
        user,
        user_id,
        data.get("Паков", 0),
        data.get("Вес", 0),
        data.get("Пакетосварка", 0),
        data.get("Флекса", 0),
        clean_extrusion(data.get("Экструзия", "")),
        data.get("Итого", 0),
    ])

    # Добавляем строку "ИТОГО"
    if sheet.max_row > 1:
        totals = ["ИТОГО", "", ""]
        for col in range(4, 9):  # Столбцы D–H (индекс 3–7)
            col_letter = chr(64 + col)
            formula = f"=SUM({col_letter}2:{col_letter}{sheet.max_row})"
            totals.append(formula)
        sheet.append(totals)

    wb.save(path)

def export_excel(folder: str) -> str:
    path = get_excel_path(folder)
    return path if os.path.isfile(path) else None

def reset_month_data(folder: str):
    path = get_excel_path(folder)
    if os.path.exists(path):
        os.remove(path)

def get_user_stats(folder: str) -> str:
    path = get_excel_path(folder)
    if not os.path.exists(path):
        return "Нет данных."

    wb = load_workbook(path)
    sheet = wb.active
    stats = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if str(row[0]).lower() == "итого":
            continue
        name = row[1]
        pakov = row[3] or 0
        ves = row[4] or 0
        otxody = (row[5] or 0) + (row[6] or 0)
        if name not in stats:
            stats[name] = {"паков": 0, "вес": 0, "отходы": 0}
        stats[name]["паков"] += pakov
        stats[name]["вес"] += ves
        stats[name]["отходы"] += otxody

    msg = "📊 <b>Статистика по пользователям:</b>
"
    for user, d in stats.items():
        msg += (
            f"
👤 <b>{user}</b>
"
            f"📦 Паков: <b>{d['паков']}</b>
"
            f"⚖️ Вес: <b>{d['вес']} кг</b>
"
            f"♻️ Отходы: <b>{d['отходы']} кг</b>
"
        )
    return msg
