import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

DATA_DIR = "/config/bnk_bot/data"
os.makedirs(DATA_DIR, exist_ok=True)

def get_file_path():
    return os.path.join(DATA_DIR, f"{datetime.now().strftime('%Y-%m')}.xlsx")

def save_entry(date, user, values):
    file_path = get_file_path()
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Дата", "Имя", "Паков", "Вес", "Пакетосварка", "Флекса", "Экструзия", "Итого"])

    ws.append([
        date.strftime('%Y-%m-%d %H:%M'),
        user,
        values.get("Паков", 0),
        values.get("Вес", 0),
        values.get("Пакетосварка", 0),
        values.get("Флекса", 0),
        values.get("Экструзия", ""),
        values.get("Итого", 0),
    ])
    wb.save(file_path)

def get_csv_file():
    return get_file_path()

def generate_stats(stats):
    lines = ["📊 Статистика по пользователям:"]
    for user, data in stats.items():
        lines.append(f"{user}: 🧃{data['Паков']} 🏋️‍♂️{data['Вес']} ❌{data['Пакетосварка']} 🎨{data['Флекса']} 📦{data['Итого']}")
    return "\n".join(lines)